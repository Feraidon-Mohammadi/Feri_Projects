import os
import openpyxl
import pandas as pd



# os.getcwd()
# os.chdir(path="produceSales.xlsx")  # this changes our CWD, if the Excel sheet is not in CWD
#



# Load the Excel file
file = "produceSales.xlsx"
data = pd.ExcelFile(file)

# Read the data from the specified sheet
df = data.parse("Sheet1")


# Display the first 10 rows of the DataFrame
print(df.head(10))


"""
Read in the spreadsheet data
The next step is to read in data from the spreadsheet [Sheet1].
"""
ps = openpyxl.load_workbook("produceSales.xlsx")



# Select the sheet
sheet = ps["Sheet1"]

# Get the maximum number of rows
max_rows = sheet.max_row
print("Total number of rows:", max_rows)

##############################################################

for row in range(2, sheet.max_row + 1):
# each row in the spreadsheet represents information for a particular purchase.
	produce = sheet["B" + str(row)].value
	cost_per_pound = sheet["C" + str(row)].value
	pounds_sold = sheet["D" + str(row)].value
	total_sales = sheet["E" + str(row)].value
# the first column is B followed by C and so on.
# Each value in a cell is represented by a column letter and a row number. So #the first element in the sheet is B1, next column C1 and so on. This enables #to iterate over the entire cells.





